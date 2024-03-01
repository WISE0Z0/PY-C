import openpyxl
book = openpyxl.load_workbook("C:\\Users\\HP\\Desktop\\PY\\Book1.xlsx")
sheet = book.active

from openpyxl import load_workbook

#Load_workbook
workbook = load_workbook("C:\\Users\\HP\\Desktop\\PY\\Book1.xlsx")

#Extract data from sheet1-
sheet1 = workbook['sheet1']
data_sheet1 = []

for row in sheet1.iter_rows(values_only=True):
    data_row = []
    for cell in row:
        data_row.append(cell.value)
    data_sheet1.append(row)

#Extract data from sheet2-
sheet2 = workbook['sheet2']
data_sheet2 = []

for row in  sheet2.iter_rows(values_only=True):
    data_sheet2.append(row)

#print the data from the sheet1-
print('data from sheet1: ')
for row in data_sheet2:
    print(row)